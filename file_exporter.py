"""
file_exporter.py
----------------
Módulo para exportar resultados a archivos CSV y Excel.

Responsabilidades:
    - Recibir un DataFrame de pandas.
    - Guardar en formato CSV o Excel en la carpeta 'outputs/'.
    - Retornar la ruta del archivo generado.

¿Por qué separar esto en un módulo propio?
    - Principio de responsabilidad única (SRP).
    - Facilita agregar nuevos formatos en el futuro (JSON, PDF, etc.).
    - Hace el código más testeable.
"""

import pandas as pd
from pathlib import Path
from datetime import datetime


OUTPUT_DIR = Path(__file__).parent / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)


def export_to_csv(df: pd.DataFrame, nombre: str) -> str:
    """
    Exporta un DataFrame a un archivo CSV.

    Parámetros:
        df:     DataFrame con los datos a exportar.
        nombre: Nombre base del archivo (sin extensión).

    Retorna:
        Ruta del archivo CSV generado.

    Formato del archivo:
        - Separador: coma (,)
        - Encoding: UTF-8 con BOM (para compatibilidad con Excel en Windows)
        - Sin índice de pandas
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    # Limpiar nombre: eliminar caracteres especiales
    safe_name = "".join(c if c.isalnum() or c in ('_', '-') else '_' for c in nombre)
    filename = OUTPUT_DIR / f"{safe_name}_{timestamp}.csv"

    # utf-8-sig agrega BOM para que Excel lo abra correctamente
    df.to_csv(filename, index=False, encoding="utf-8-sig")

    print(f"📄 CSV exportado: {filename} ({len(df)} filas, {len(df.columns)} columnas)")
    return str(filename)


def export_to_excel(df: pd.DataFrame, nombre: str) -> str:
    """
    Exporta un DataFrame a un archivo Excel (.xlsx) con formato básico.

    Parámetros:
        df:     DataFrame con los datos a exportar.
        nombre: Nombre base del archivo (sin extensión).

    Retorna:
        Ruta del archivo Excel generado.

    Características del Excel generado:
        - Encabezados en negrita con fondo azul claro.
        - Columnas con ancho ajustado al contenido.
        - Hoja llamada 'Ventas'.
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = "".join(c if c.isalnum() or c in ('_', '-') else '_' for c in nombre)
    filename = OUTPUT_DIR / f"{safe_name}_{timestamp}.xlsx"

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Ventas")

        # ── Dar formato al Excel ──────────────────────
        workbook = writer.book
        worksheet = writer.sheets["Ventas"]

        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")

        # Formato de encabezados
        for col_num, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Ancho automático de columnas
        for col_num, col_name in enumerate(df.columns, 1):
            max_len = max(len(str(col_name)), df[col_name].astype(str).str.len().max())
            worksheet.column_dimensions[get_column_letter(col_num)].width = min(max_len + 4, 40)

    print(f"📊 Excel exportado: {filename} ({len(df)} filas, {len(df.columns)} columnas)")
    return str(filename)
